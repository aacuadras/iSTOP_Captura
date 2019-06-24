import pandas as pd
import csv, pyodbc
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

#Database reader initialization
MDB = 'c:/SetupPV/Data/DBTRUCK.mdb' 
DRV = '{Microsoft Access Driver (*.mdb)}'
PWD = 'pw'
#Driver initialization
con = pyodbc.connect('DRIVER={};DBQ={};PWD={}'.format(DRV,MDB,PWD))
cur = con.cursor()
#Workbook loading
wb = load_workbook('./master_folios.xlsx')
sheet = wb['MASTER']
print("Leyendo hoja: " + sheet.title)


#This loop will iterate through all the rows of the datasheet, the program must run an sql query to update the database with
#the according data
for i in range(4, sheet.max_row):
      if(sheet['G' + str(i)].value == 'TOTAL'):
            break
      
      #Update query
      SQL = "UPDATE HISTORIA SET guia='" + str(sheet['H' + str(i)].value) + "', FOLIOGEN='" + str(sheet['D' + str(i)].value) + "', PROVEEDOR='" + str(sheet['C' + str(i)].value) + "' WHERE PLACAS='" + str(sheet['G' + str(i)].value) + "' AND FECHA1='" \
      + str(sheet['B' + str(i)].value.strftime('%d/%m/%Y')) + "' AND observacion ='" + str(sheet['F' + str(i)].value) + "';"
      cur.execute(SQL)
      cur.commit()
      

print('Base de datos actualizada')